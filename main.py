import tkinter as tk
import datetime
import openpyxl
import tkinter.messagebox as messagebox
import pandas as pd
import os.path

# Check if Excel file exists
if os.path.isfile("D:/worktimer.xlsx"):
    # Load data from Excel workbook into a Pandas DataFrame
    df = pd.read_excel("D:/worktimer.xlsx")
else:
    # Create new Excel workbook with necessary columns
    df = pd.DataFrame(columns=['Start Time', 'End Time', 'Total Time'])
    df.to_excel("D:/worktimer.xlsx", index=False)

# Group data by date and calculate total time worked
df['Start Time'] = pd.to_datetime(df['Start Time'])
df['End Time'] = pd.to_datetime(df['End Time'])
df['Date'] = df['Start Time'].dt.date
grouped = df.groupby(['Date'])['Total Time'].sum().reset_index()



# Create GUI form
root = tk.Tk()
root.geometry("350x250")
root.title("Work Statistic")
root.config(bg='#464646')

# Define font for title
title_font = ("Arial", 14, "bold")


def set_start_time():
    # Get current time and format as string
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Set start time entry to current time
    start_entry.delete(0, tk.END)
    start_entry.insert(0, now)

    # Change color of Start button
    start_button.config(bg='#68DD00')
    end_button.config(bg=root['bg'])


def set_end_time():
    # Get current time and format as string
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Set end time entry to current time
    end_entry.delete(0, tk.END)
    end_entry.insert(0, now)

    # Change color of End button
    end_button.config(bg='#EC0000')
    start_button.config(bg=root['bg'])


start_label = tk.Label(root, font=title_font, fg='#FFFFFF', bg='#464646')
start_label.pack()
start_entry = tk.Entry(root)
start_entry.pack()
start_label.pack(side=tk.LEFT, padx=(10, 0), pady=(10, 10))
start_button_frame = tk.Frame(root, bg=root['bg'])
start_button_frame.pack(pady=20)
start_button = tk.Button(start_button_frame, text="Start Work", command=set_start_time, bg=root['bg'], fg='#FFFFFF')
start_button.pack(side=tk.LEFT)

end_label = tk.Label(root, font=title_font, fg='#FFFFFF', bg='#464646')
end_label.pack()
end_entry = tk.Entry(root)
end_entry.pack()

end_button_frame = tk.Frame(root, bg=root['bg'])
end_button_frame.pack(pady=20)
end_button = tk.Button(end_button_frame, text="Finish Work", command=set_end_time, bg=root['bg'], fg='#FFFFFF')
end_button.pack(side=tk.LEFT)


def calculate():
    # Get start and end time from GUI form
    start_time_str = start_entry.get()
    start_time = datetime.datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")

    end_time_str = end_entry.get()
    end_time = datetime.datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # Calculate total time worked
    total_time = end_time - start_time
    total_time_str = str(total_time).split(".")[0]  # Convert total time to HH:MM:SS format

    # Write results to Excel file
    wb = openpyxl.load_workbook('D:/worktimer.xlsx')
    sheet = wb.active

    # set column names
    sheet.cell(row=1, column=1).value = "Start Time"
    sheet.cell(row=1, column=2).value = "End Time"
    sheet.cell(row=1, column=3).value = "Total Time"

    # append new entry
    row = [start_time.strftime("%Y-%m-%d %H:%M:%S"), end_time.strftime("%Y-%m-%d %H:%M:%S"), total_time_str]
    sheet.append(row)
    wb.save('D:/worktimer.xlsx')

    # Show message box with total time worked
    tk.messagebox.showinfo("Total Time Worked", f"You worked for {total_time_str}")


calculate_button = tk.Button(root, text="Calculate", command=calculate)
calculate_button.pack()


root.mainloop()
